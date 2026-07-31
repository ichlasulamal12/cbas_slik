"""
==============================================================================
Project : CBAS SLIK Feature Engineering
File    : report.py
Author  : Ichlasul Amal
Version : 1.0.0
==============================================================================

Feature Engineering Report

Generate QA Report dalam satu file Excel.

"""

from pathlib import Path
from datetime import datetime

import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Font


# =============================================================================
# WRITE DATAFRAME
# =============================================================================

def write_dataframe(ws, df: pl.DataFrame):

    if df is None:
        return

    ws.append(df.columns)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in df.iter_rows():
        ws.append(list(row))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# =============================================================================
# SUMMARY
# =============================================================================

def create_summary(df, catalog, pipeline_info=None):

    numeric = sum(
        "Int" in str(t) or "Float" in str(t)
        for t in df.schema.values()
    )

    categorical = sum(
        str(t) == "String"
        for t in df.schema.values()
    )

    summary = {

        "Metric":[

            "Generated At",
            "Total Records",
            "Total Features",
            "Numeric Features",
            "Categorical Features",
            "Feature Registry",
            "Project",
            "Pipeline Version",

        ],

        "Value":[

            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            str(df.height),
            str(df.width),
            str(numeric),
            str(categorical),
            str(catalog.height),
            str(
                pipeline_info.get("project", "CBAS SLIK")
                if pipeline_info
                else "CBAS SLIK"
            ),
            str(
                pipeline_info.get("version", "1.0.0")
                if pipeline_info
                else "1.0.0"
            ),

        ]

    }

    return pl.DataFrame(summary)


# =============================================================================
# MODULE
# =============================================================================

def create_module_summary(catalog):

    return (

        catalog

        .group_by("module")

        .len()

        .rename({"len":"feature_count"})

        .sort("feature_count", descending=True)

    )


# =============================================================================
# DOMAIN
# =============================================================================

def create_domain_summary(catalog):

    return (

        catalog

        .group_by("domain")

        .len()

        .rename({"len":"feature_count"})

    )


# =============================================================================
# CATEGORY
# =============================================================================

def create_category_summary(catalog):

    return (

        catalog

        .group_by("category")

        .len()

        .rename({"len":"feature_count"})

    )


# =============================================================================
# LEVEL
# =============================================================================

def create_level_summary(catalog):

    return (

        catalog

        .group_by("level")

        .len()

        .rename({"len":"feature_count"})

    )


# =============================================================================
# MISSING
# =============================================================================

def create_missing_summary(catalog):

    return (

        catalog

        .select(

            [

                "feature_name",
                "missing_count",
                "missing_ratio"

            ]

        )

        .sort(

            "missing_ratio",

            descending=True

        )

    )


# =============================================================================
# CONSTANT FEATURE
# =============================================================================

def create_constant_feature(catalog):

    return (

        catalog

        .filter(

            pl.col("unique_count") <= 1

        )

    )


# =============================================================================
# FEATURE DEPENDENCY
# =============================================================================

def create_dependency_report(catalog):

    cols = [

        c

        for c in [

            "feature_name",
            "module",
            "domain",
            "level",
            "formula",
            "source_columns",
            "dependencies"

        ]

        if c in catalog.columns

    ]

    return catalog.select(cols)


# =============================================================================
# PIPELINE EXECUTION
# =============================================================================

def create_pipeline_report(pipeline_info):

    if pipeline_info is None:

        return pl.DataFrame(

            {

                "Metric":[

                    "Status"

                ],

                "Value":[

                    "Pipeline information not provided"

                ]

            }

        )

    rows=[]

    for key,value in pipeline_info.items():

        rows.append(

            {

                "Metric":key,

                "Value":str(value)

            }

        )

    return pl.DataFrame(rows)


# =============================================================================
# MAIN
# =============================================================================

def create_feature_report(

    df,

    catalog,

    output_dir,

    pipeline_info=None,

):

    print("="*80)
    print("Create Feature Report")
    print("="*80)

    wb = Workbook()

    wb.remove(wb.active)

    sheets = {

        "Summary":

            create_summary(

                df,

                catalog,

                pipeline_info,

            ),

        "Feature Catalog":

            catalog,

        "Module Summary":

            create_module_summary(

                catalog

            ),

        "Domain Summary":

            create_domain_summary(

                catalog

            ),

        "Category Summary":

            create_category_summary(

                catalog

            ),

        "Level Summary":

            create_level_summary(

                catalog

            ),

        "Missing Summary":

            create_missing_summary(

                catalog

            ),

        "Constant Feature":

            create_constant_feature(

                catalog

            ),

        "Feature Dependency":

            create_dependency_report(

                catalog

            ),

        "Pipeline Execution":

            create_pipeline_report(

                pipeline_info

            ),

    }

    for sheet_name, dataframe in sheets.items():

        ws = wb.create_sheet(sheet_name)

        write_dataframe(

            ws,

            dataframe,

        )

    output_file = Path(output_dir) / "feature_report.xlsx"

    wb.save(output_file)

    print(f"Report saved : {output_file}")

    return output_file
